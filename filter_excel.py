import functools
import json
import logging
import operator
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame

from strategies import calculation_registry

logger = logging.getLogger(__name__)


def load_folder(folder: str = r"./source") -> list:
    """加载目录"""

    file_folder = Path(folder)

    if not file_folder.exists():
        logger.error(f"路径不存在,请进行检查：{file_folder}")
        return []
    logger.info(f"正在加载路径目录：{folder}")
    exist_suffix = [".xlsx", ".xls"]

    excel_files_path = [
        file for file in file_folder.rglob("*") if file.suffix in exist_suffix
    ]
    if not excel_files_path:
        logger.warning(f"{file_folder}中没有找到excel文件")
        return []

    logger.info(f"共找到{len(excel_files_path)}个excel文件,准备处理....")

    return excel_files_path


def load_all_sheet(file_path: str) -> dict[str, DataFrame]:
    """加载所有sheet"""
    dfs = pd.read_excel(file_path, engine="calamine", sheet_name=None)
    return dfs


def init_config() -> dict:
    """初始化config"""
    with open("./config/config.json", "r", encoding="utf-8") as f:
        return json.load(f)


def get_keywords(keywords_path: str) -> str:
    """加载关键字，并进行转换"""
    content = ""
    with open(keywords_path, mode="r", encoding="utf-8") as f:
        content = f.read()
    return "|".join(content.strip().split())


def search_only(df: DataFrame, keywords: str, filter_columns: list) -> DataFrame:
    """查询关键字"""
    mask_list = []
    for column in filter_columns:
        if column in df.columns:
            result = (
                df[column]
                .fillna("")
                .astype(str)
                .str.contains(keywords, na=False, case=False)
            )
            mask_list.append(result)
    final_mask = functools.reduce(operator.or_, mask_list)

    return df[final_mask]


def search_add_keywords(
    df: DataFrame, keywords: str, filter_columns: list
) -> DataFrame:
    """查询关键字，并且添加匹配到的关键字"""
    match_df = search_only(df, keywords, filter_columns)

    if match_df.empty:
        return match_df

    all_match_list = pd.Series([[] for _ in range(len(match_df))], index=match_df.index)

    for column in filter_columns:
        if column in match_df.columns:
            all_match_list += match_df[column].fillna("").str.findall(keywords)

    def format_keywords(keyword_list: list) -> str:
        """格式化关键字列"""
        unique_keywords = {k.lower() for k in keyword_list}
        return " ".join(sorted(list(unique_keywords)))

    match_df["关键字"] = all_match_list.apply(format_keywords)
    return match_df


def save_excel(df: DataFrame) -> None:
    """使用模板进行保存"""
    # 加载模板
    wb = load_workbook("./template/template.xlsx")
    ws = wb["Sheet1"]

    start_row = 2
    rows_to_write = dataframe_to_rows(df, index=False, header=False)
    for r_idx, row_values in enumerate(rows_to_write, start=start_row):
        for c_idx, value in enumerate(row_values, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save("./output/结果.xlsx")


def extract_statistics(
    df: DataFrame, calculations: list[dict[str, Any]], profile_config: dict[str, Any]
) -> dict[str, Any]:
    """
    统计dataframe中的数据，返回字典

    df: dataframe,
    calculations:
    profile_config:
    """

    if df.empty:
        raise ValueError("df不能为空")

    stats_data = {}
    column_mapping = profile_config.get("column_mapping", {})

    for calc in calculations:
        key = calc.get("key")
        calc_type = calc.get("type", "")

        if not all([key, calc_type]):
            logger.warning("警告：当前calculation,缺少'key'或'type'。进行跳过")
            continue

        strategy = calculation_registry.get(calc_type)
        if not strategy:
            logger.warning(f"calculation计算项 '{calc_type}'，跳过 '{key}'.")
            stats_data[key] = f"未知类型{calc_type}"
            continue

        actual_column_name = None
        if "column_ref" in calc:
            logical_name = calc["column_ref"]
            actual_column_name = column_mapping.get(logical_name)
            if not actual_column_name:
                logger.warning(
                    f"在Profile中'{profile_config.get('name')}'找不到逻辑列'{logical_name}'的映射，跳过'{key}'。"
                )
                stats_data[key] = f"列映射缺失{logical_name}"
                continue

        params = {"df": df, "column": actual_column_name, "value": calc.get("value")}

        try:
            result = strategy.execute(**params)
            stats_data[key] = result
        except Exception as e:
            logger.error(f"错误：执行'{key}'({calc_type})时发生错误:{e}")
            stats_data[key] = "计算错误"

    return stats_data


def format_statistics_text(
    stats_data: dict[str, Any], format_config: dict[str, Any]
) -> str:
    """
    根据extract_statistics生成是数据字典，生成报告文本

    stats: 统计后的数据字典
    """

    template = format_config.get("template", "文字模板未在配置文件中找到")

    format_placeholers = {}

    for key, result in stats_data.items():
        if result is None:
            format_placeholers[key] = "无效数据"
        elif isinstance(result, dict) and "start" in result and "end" in result:
            date_format_str = format_config.get("date", "%Y-%m-%d")
            start_str = result["start"].strftime(date_format_str)
            end_str = result["end"].strftime(date_format_str)
            format_placeholers[key] = f"{start_str}至{end_str}"
        elif isinstance(result, dict):
            item_unit = format_config.get("item_unit", "")
            separator = format_config.get("item_separator", ",")
            if not result:
                format_placeholers[key] = "无相关信息"
            else:
                parts = [f"{name}{value}{item_unit}" for name, value in result.items()]
                format_placeholers[key] = separator.join(parts)
        else:
            format_placeholers[key] = result

    try:
        final_text = template.format(**format_placeholers)
    except KeyError as e:
        return f"模板格式化错误：模板中的占位符'{e}'在计算中找到对应的值"

    return final_text


def load_and_merge_sheets(sheets: dict[str, DataFrame]) -> DataFrame | None:
    """合并所有sheet,进行后面统计"""

    if not sheets:
        logger.warning("统计合并Dataframe为空")
        return None

    sheet_columns = {
        name: tuple(sorted(df.columns.tolist()))
        for name, df in sheets.items()
        if df is not None and not df.empty
    }

    compatible_sheets = []
    incompatible_sheets = []

    first_sheet_name = next(iter(sheet_columns))
    primary_columns = sheet_columns[first_sheet_name]

    for name, columns in sheet_columns.items():
        if columns == primary_columns:
            compatible_sheets.append(name)
        else:
            incompatible_sheets.append(name)

    dfs_to_concat = [sheets[name] for name in compatible_sheets]

    try:
        merged_df = pd.concat(dfs_to_concat, ignore_index=True)
        logger.info(
            f"统计：合并了{len(compatible_sheets)}个工作表，{', '.join(compatible_sheets)}"
        )
        return merged_df
    except Exception as e:
        logger.error(f"统计：合并工作表时发生错误：{e}")
        return None


def main():
    pass


if __name__ == "__main__":
    main()
